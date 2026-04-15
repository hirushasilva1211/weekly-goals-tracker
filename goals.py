import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import load_workbook
import os

FILE = "weekly_goals.xlsx"

def get_or_create_workbook():
    if os.path.exists(FILE):
        return load_workbook(FILE)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Goals"

    # Define thick border
    thick = Side(style="thick")
    thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    # Define background color (steel blue)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Write headers with bold font + thick border + background color
    headers = ["Week", "Goal", "Status"]
    for col, header in enumerate(headers, 2):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")  # white text
        cell.border = thick_border
        cell.fill = header_fill

    wb.save(FILE)
    return wb

def apply_border_to_merged_cell(ws, start_row, end_row, col):
    thin = Side(style="thin")
    none = Side(style=None)

    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=col)

        top    = thin if row == start_row else none
        bottom = thin if row == end_row   else none

        # Left and right always apply (they're the outer edges)
        cell.border = Border(
            left=thin,
            right=thin,
            top=top,
            bottom=bottom
        )

def add_week(week_label, goals):
    wb = get_or_create_workbook()
    ws = wb["Goals"]

    start_row = ws.max_row + 2  # Next empty row after existing data

    # Define thin border
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, goal in enumerate(goals):
        row = start_row + i
        cell1 = ws.cell(row=row, column=3, value=goal)
        cell2 = ws.cell(row=row, column=4, value="")
        cell1.border = thin_border
        cell2.border = thin_border

    # Merge the "Week" column across all goal rows
    end_row = start_row + len(goals) - 1
    ws.merge_cells(f"B{start_row}:B{end_row}")

    # Write the week label into the merged cell
    merged_cell = ws.cell(row=start_row, column=2, value=week_label)
    merged_cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    merged_cell.font = Font(bold=True)
    apply_border_to_merged_cell(ws, start_row, end_row, col=2)

    try:
        wb.save(FILE)
        print(f"✅ Saved {len(goals)} goals for {week_label}")
    except:
        print("Something Went Wrong!")


# --- Usage ---
week = "Week 02\nApr 13 - Apr 19"
goals = [
    "Learn Docker basics",
    "Complete internship task",
    "Read 15 pages",
    "Review Angular signals",
]
input("⚠️  Make sure weekly_goals.xlsx is closed in Excel. Press Enter to continue...")
add_week(week, goals)